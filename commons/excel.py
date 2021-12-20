from openpyxl import Workbook, load_workbook
import os

from config.publicconfig import RESULT_PATH

def category(categoryNo=20, types=[]):
    categoryList = []
    for i in range(categoryNo):
        categoryList.append(i+1)
    if len(types) == 0:
        newCategoryList = []
    elif len(types) == 1:
        newCategoryList = categoryList
    else:
        newCategoryList = []
        for i in range(len(types) - 1):
            newCategoryList = newCategoryList + categoryList
    return newCategoryList

# https://goodthings4me.tistory.com/487
def writeInExcel(dataList, dataType='data', fileName='분석 결과', sheetName='결과'):
    '''
        1. file이 있는지 확인 - 파일이 있다면 load 없으면 새로운 workbook 생성
        2. sheet가 있는지 확인 - 없으면 새로운 sheet 생성
        3. type 확인
            -> title 인 경우 1 row(행)에 2번째 cloumn(열) 부터 순서래도 data write
            -> data 인 경우 dataList를 행단위로 추가
    '''
    if os.path.exists(RESULT_PATH):
        pass
    else:
        os.mkdir(RESULT_PATH)
    filePath = os.path.join(RESULT_PATH, fileName+'.xlsx')

    if os.path.isfile(filePath):
        wb = load_workbook(filePath)
    else:
        wb = Workbook()

    # 이름이 있는 시트를 생성
    try:
        ws = wb[sheetName]
    except Exception as e:
        print(e)
        ws = wb.create_sheet(sheetName)

    # 행 단위로 추가
    if dataType == 'title':
        for i, data in enumerate(dataList):
            ws.cell(1, i+2, data)
    else:
        ws.append(dataList)
    wb.save(filePath)

def summarize(categoryNo=20, fileName='분석 결과', sheetName='결과'):
    '''
        1. data 수집해 있는 sheet를 불러 들임
        2. 시간대 별로 data 확인
        3. 시간대가 내림차순으로 되어 있어 오름 차순으로 변경하기 위해서 reverse 진행
        4. 누적 data 계산
        5. new sheet에 data 기록
    '''
    filePath = os.path.join(RESULT_PATH, fileName+'.xlsx')
    wb = load_workbook(filePath)
    ws = wb[sheetName]
    # 이름이 있는 시트를 생성
    maxRow = ws.max_row
    dateList = []
    dataDict = {}

    for i in range(maxRow - 1):
        dateList.append(ws.cell(i + 2, 1).value)
        dataDict[ws.cell(i+2, 1).value] = []
        for j in range(categoryNo):
            result = ws.cell(i + 2, categoryNo + 2 + j).value
            dataDict[ws.cell(i+2, 1).value].append(result)

    dateList.reverse()
    newDataDict = {}
    lastDate = 0
    for i, date in enumerate(dateList):
        newDataDict[date] = []
        for j in range(categoryNo):
            if i == 0:
                result = dataDict[date][j]
                newDataDict[date].append(result)
            else:
                result = round((newDataDict[lastDate][j] + 1) * (dataDict[date][j] + 1) - 1, 4)
                newDataDict[date].append(result)
        lastDate = date

    sheetName = 'sm_' + sheetName
    titleList = category(categoryNo=20, types=['newResult'])
    writeInExcel(titleList, dataType='title', sheetName=sheetName)
    for date in dateList:
        resultList = [date] + newDataDict[date]
        writeInExcel(resultList, dataType='data', sheetName=sheetName)



